# api/main.py
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import sqlite3
import os
import sys
from storage.database import init_db
from config import DB_PATH
import json

CONFIG_PATH = "data/config.json"
DEFAULT_CONFIG = {
    "keywords":           ["Product Manager", "Associate Product Manager"],
    "locations":          ["Bengaluru", "Hyderabad"],
    "match_threshold":    75,
    "outreach_threshold": 75,
}

app = FastAPI(title="COIE API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],   # Tighten in production
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
def startup():
    init_db()

def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def read_config() -> dict:
    if os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH) as f:
            return json.load(f)
    return DEFAULT_CONFIG.copy()

def write_config(cfg: dict):
    os.makedirs("data", exist_ok=True)
    with open(CONFIG_PATH, "w") as f:
        json.dump(cfg, f, indent=2)

@app.get("/api/config")
def get_config():
    return read_config()

@app.patch("/api/config")
def update_config(body: dict):
    cfg = read_config()
    # Only update known keys — ignore anything else
    for key in ["keywords", "locations", "match_threshold", "outreach_threshold"]:
        if key in body:
            cfg[key] = body[key]
    write_config(cfg)
    return cfg

@app.get("/api/jobs")
def get_jobs(min_score: float = 0, limit: int = 100):
    """Return all jobs sorted by match score."""
    with get_conn() as conn:
       rows = conn.execute("""
            SELECT hash_id, title, company, location, source,
                   url, posted_text, scraped_at, match_score,
                   recruiter, recruiter_title, recruiter_email,
                   recruiter_linkedin, email_subject, email_body, status
            FROM jobs
            WHERE match_score >= ?
            AND status NOT IN ('Skipped', 'Rejected')  
            ORDER BY match_score DESC
            LIMIT ?
        """, (min_score, limit)).fetchall()
    return [dict(r) for r in rows]

@app.get("/api/jobs/skipped")
def get_skipped_jobs():
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT hash_id, title, company, location, source,
                   url, posted_text, scraped_at, match_score,
                   recruiter, recruiter_title, recruiter_email,
                   recruiter_linkedin, email_subject, email_body, status
            FROM jobs
            WHERE status = 'Skipped'
            ORDER BY match_score DESC
        """).fetchall()
    return [dict(r) for r in rows]


@app.get("/api/stats")
def get_stats():
    """Return pipeline summary stats for the dashboard."""
    with get_conn() as conn:
        total = conn.execute("SELECT COUNT(*) FROM jobs").fetchone()[0]
        high  = conn.execute(
            "SELECT COUNT(*) FROM jobs WHERE match_score >= ?",
            (read_config()["match_threshold"],)
        ).fetchone()[0]
        # Count jobs that ever reached "Outreach Sent" or beyond
        sent = conn.execute("""
            SELECT COUNT(*) FROM jobs 
            WHERE status IN ('Outreach Sent', 'Replied', 'Interview')
        """).fetchone()[0]
        # Count jobs that ever reached "Replied" or beyond
        replied = conn.execute("""
            SELECT COUNT(*) FROM jobs 
            WHERE status IN ('Replied', 'Interview')
        """).fetchone()[0]
        by_source = conn.execute("""
            SELECT source, COUNT(*) as count
            FROM jobs GROUP BY source
        """).fetchall()
        top_jobs = conn.execute("""
            SELECT title, company, match_score, source, status
            FROM jobs
            WHERE match_score >= ?
            AND status NOT IN ('Skipped', 'Rejected')
            ORDER BY match_score DESC
            LIMIT 5
        """,(read_config()["match_threshold"],)).fetchall()

    return {
        "total_scraped":        total,
        "high_match":           high,
        "outreach_sent":        sent,
        "replies":              replied,
        "threshold":            read_config()["match_threshold"],
        "outreach_threshold":   read_config()["outreach_threshold"],
        "reply_rate":           round((replied / sent * 100) if sent > 0 else 0, 1),
        "by_source":            [dict(r) for r in by_source],
        "top_jobs":             [dict(r) for r in top_jobs],
    }


@app.patch("/api/jobs/{hash_id}/status")
def update_status(hash_id: str, body: dict):
    """Update job status (e.g. Outreach Sent, Replied)."""
    status = body.get("status")
    if not status:
        return {"error": "status required"}
    with get_conn() as conn:
        conn.execute(
            "UPDATE jobs SET status = ? WHERE hash_id = ?",
            (status, hash_id)
        )
    return {"ok": True, "hash_id": hash_id, "status": status}


@app.patch("/api/jobs/{hash_id}/recruiter")
def update_recruiter(hash_id: str, body: dict):
    """Manually add recruiter name + email."""
    with get_conn() as conn:
        conn.execute(
            "UPDATE jobs SET recruiter = ?, email = ? WHERE hash_id = ?",
            (body.get("recruiter", ""), body.get("email", ""), hash_id)
        )
    return {"ok": True}

@app.post("/api/jobs/{hash_id}/outreach")
def perform_outreach(hash_id: str):
    """Trigger Hunter + Groq for a specific job on demand."""
    from recruiter.hunter   import find_recruiter
    from outreach.generator import generate_email
    from storage.database   import update_outreach

    # Fetch the job from DB
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM jobs WHERE hash_id = ?", (hash_id,)
        ).fetchone()

    if not row:
        return {"error": "Job not found"}

    job = dict(row)

    if not job.get("description") or len(job["description"]) < 50:
        return {"error": "No job description available to personalize email"}

    # Hunter recruiter lookup
    recruiter = find_recruiter(job["company"], job["location"]) or {}

    # Groq email generation
    email = generate_email(
        job_title       = job["title"],
        company         = job["company"],
        job_description = job["description"],
        recruiter       = recruiter,
    )

    if not email:
        return {"error": "Email generation failed"}

    update_outreach(hash_id, recruiter, email)

    return {
        "ok":               True,
        "recruiter":        recruiter.get("name", ""),
        "recruiter_email":  recruiter.get("email", ""),
        "email_subject":    email["subject"],
        "email_body":       email["body"],
    }

@app.get("/api/jobs/export")
def export_jobs():
    cfg = read_config()
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT title, company, location, source, match_score,
                   posted_text, status, url, recruiter, recruiter_email
            FROM jobs
            WHERE match_score > 0
            AND status NOT IN ('Skipped', 'Rejected')
            ORDER BY match_score DESC
        """).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "COIE Job Feed"

    # Header style
    header_fill = PatternFill("solid", fgColor="1B4F72")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    headers = ["Title", "Company", "Location", "Source", "Match Score",
               "Posted", "Status", "Recruiter", "Recruiter Email", "URL"]
    col_widths = [35, 25, 20, 12, 14, 14, 16, 22, 30, 50]

    for i, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 20

    # Score color coding
    def score_fill(score):
        if score >= 80:   return PatternFill("solid", fgColor="D5F5E3")  # green
        elif score >= 75: return PatternFill("solid", fgColor="FEF9E7")  # yellow
        else:             return PatternFill("solid", fgColor="F2F3F4")  # gray

    # Data rows
    for ri, row in enumerate(rows, 2):
        values = [
            row["title"],
            row["company"],
            row["location"],
            row["source"],
            f"{row['match_score']}%",
            row["posted_text"] or "—",
            row["status"],
            row["recruiter"] or "—",
            row["recruiter_email"] or "—",
            row["url"],
        ]
        fill = score_fill(row["match_score"])
        for ci, value in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            # Color code score column and row
            if ci == 5:
                cell.fill = fill
                cell.font = Font(name="Calibri", size=10, bold=True)

        ws.row_dimensions[ri].height = 16

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto filter
    ws.auto_filter.ref = f"A1:J{len(rows)+1}"

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    from datetime import datetime
    filename = f"COIE_Jobs_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

from fastapi.staticfiles import StaticFiles
app.mount("/", StaticFiles(directory="dashboard/dist", html=True), name="static")